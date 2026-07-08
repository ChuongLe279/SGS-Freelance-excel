import json
import re
import sys
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path

import duckdb
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz, process

try:
    from langchain_google_genai import ChatGoogleGenerativeAI
except ImportError:
    ChatGoogleGenerativeAI = None

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

try:
    from setup import (
        bank_aliases,
        banks,
        province_old_to_new_alias,
        provinces,
        ward_old_to_new_alias,
        wards,
    )
except ImportError:
    bank_aliases = {}
    banks = []
    province_old_to_new_alias = {}
    provinces = []
    ward_old_to_new_alias = {}
    wards = []


def remove_diacritics(text):
    text = unicodedata.normalize("NFD", str(text))
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = text.replace("đ", "d").replace("Đ", "D")
    return text


def normalize(s):
    s = remove_diacritics(str(s)).lower()
    s = re.sub(r"[^a-z0-9]+", " ", s).strip()
    return s


def normalized_column_name(name):
    return str(name).replace("\xa0", " ").strip().lower()


def find_column(df, *needles):
    normalized_needles = [normalized_column_name(needle) for needle in needles]
    for column in df.columns:
        normalized = normalized_column_name(column)
        if all(needle in normalized for needle in normalized_needles):
            return column
    raise KeyError(f"Cannot find column containing: {needles}")


@dataclass
class ColumnConfig:
    name: str
    canonical_list: list
    alias_dict: dict = field(default_factory=dict)
    threshold_high: int = 80
    threshold_fallback: int = 65
    llm_batch_size: int = 15

    def build_lookup(self):
        lookup = {}

        for canonical in self.canonical_list:
            lookup[normalize(canonical)] = canonical

        for canonical, aliases in self.alias_dict.items():
            for alias in aliases:
                lookup[normalize(alias)] = canonical

        return lookup


class ClosedSetMatcher:
    def __init__(self, config: ColumnConfig, llm_model=None):
        self.config = config
        self.lookup = config.build_lookup()
        self.lookup_keys = list(self.lookup.keys())
        self.llm_model = llm_model

    def _rule_based(self, value):
        q = normalize(value)
        if q in self.lookup:
            return self.lookup[q], 100, "tier1_rule_based"
        return None, 0, None

    def _rapidfuzz_high(self, value):
        q = normalize(value)
        if not q or not self.lookup_keys:
            return None, 0, None

        best_match = None
        best_score = 0
        scorers = [
            fuzz.token_sort_ratio,
            fuzz.token_set_ratio,
            fuzz.WRatio,
        ]

        for scorer in scorers:
            match = process.extractOne(q, self.lookup_keys, scorer=scorer)
            if not match:
                continue

            candidate, score, _ = match
            if score > best_score:
                best_score = score
                best_match = candidate

        if best_match and best_score >= self.config.threshold_high:
            return self.lookup[best_match], best_score, "tier2_rapidfuzz_high"

        return None, best_score, None

    def _rapidfuzz_fallback(self, value):
        q = normalize(value)
        if not q or not self.lookup_keys:
            return None, 0, None

        match = process.extractOne(q, self.lookup_keys, scorer=fuzz.token_set_ratio)
        if not match:
            return None, 0, None

        candidate, score, _ = match
        if score >= self.config.threshold_fallback:
            return self.lookup[candidate], score, "tier3_rapidfuzz_fallback"
        return None, score, None

    def _llm_classify(self, values):
        if not self.llm_model or not values:
            return {}

        results = {}
        canonical_list = self.config.canonical_list

        for i in range(0, len(values), self.config.llm_batch_size):
            batch = values[i:i + self.config.llm_batch_size]
            prompt = f"""
Ban la chuyen gia chuan hoa du lieu hanh chinh/ngan hang Viet Nam cho cot "{self.config.name}".
Danh sach gia tri CHUAN (closed-set, chi chon trong day):
{json.dumps(canonical_list, ensure_ascii=False, indent=2)}

Map moi gia tri "ban" sau ve dung 1 gia tri chuan. Neu khong chac chan, tra "UNKNOWN".
Gia tri can phan loai:
{json.dumps(batch, ensure_ascii=False, indent=2)}

Chi tra JSON {{"gia tri ban": "gia tri chuan hoac UNKNOWN"}}, khong giai thich, khong markdown.
"""
            try:
                response = self.llm_model.invoke(prompt)
                text = getattr(response, "content", str(response))
                text = text.strip().replace("```json", "").replace("```", "")
                results.update(json.loads(text))
            except Exception:
                for value in batch:
                    results[value] = "UNKNOWN"

        return results

    def run(self, values):
        records = [{"raw": value, "matched": None, "score": 0, "tier": None} for value in values]

        remaining = []
        for idx, record in enumerate(records):
            matched, score, tier = self._rule_based(record["raw"])
            if matched:
                record.update(matched=matched, score=score, tier=tier)
            else:
                remaining.append(idx)

        still = []
        for idx in remaining:
            matched, score, tier = self._rapidfuzz_high(records[idx]["raw"])
            if matched:
                records[idx].update(matched=matched, score=score, tier=tier)
            else:
                still.append(idx)

        still2 = []
        for idx in still:
            matched, score, tier = self._rapidfuzz_fallback(records[idx]["raw"])
            if matched:
                records[idx].update(matched=matched, score=score, tier=tier)
            else:
                still2.append(idx)

        if still2:
            names_for_llm = [records[idx]["raw"] for idx in still2]
            llm_results = self._llm_classify(names_for_llm)
            for idx in still2:
                answer = llm_results.get(records[idx]["raw"], "UNKNOWN")
                if answer in self.config.canonical_list:
                    records[idx].update(matched=answer, score=None, tier="tier4_llm")
                else:
                    records[idx].update(matched=None, score=None, tier="tier5_review_tay")

        return pd.DataFrame(records)


def build_llm_model():
    if ChatGoogleGenerativeAI is None:
        return None

    load_dotenv()
    try:
        return ChatGoogleGenerativeAI(model="gemini-2.5-flash")
    except Exception:
        return None


def is_tier2_or_below(tier_value):
    if pd.isna(tier_value):
        return False

    match = re.search(r"tier(\d+)", str(tier_value).lower())
    return bool(match and int(match.group(1)) >= 2)


def build_new_employee_lookup(new_employee_df):
    return pd.DataFrame({
        "WD ID": new_employee_df[find_column(new_employee_df, "WD ID")],
        "ID same team": new_employee_df[find_column(new_employee_df, "ID same team")],
        "Tỉnh/Thành phố": new_employee_df[find_column(new_employee_df, "Province/City")],
        "Xã/Phường": new_employee_df[find_column(new_employee_df, "Ward")],
        "Địa chỉ đường": new_employee_df[find_column(new_employee_df, "No. & Street")],
        "Số tài khoản": new_employee_df[find_column(new_employee_df, "Bank Account No.")],
        "Tên ngân hàng": new_employee_df[find_column(new_employee_df, "Bank Name")],
        "Chi nhánh ngân hàng": new_employee_df[find_column(new_employee_df, "Bank", "Branch")],
    })


def process_excel(input_path, output_path):
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    sheets = pd.read_excel(input_path, sheet_name=None)
    master_df = sheets[list(sheets.keys())[0]]
    hire_team_df = sheets[list(sheets.keys())[1]]
    new_employee_df = build_new_employee_lookup(sheets[list(sheets.keys())[2]])
    employee_list_df = sheets[list(sheets.keys())[3]]

    result = employee_list_df[["Employee Local ID"]].copy()
    result["WorkDay ID"] = employee_list_df[["WorkDay ID"]].copy()
    result["First Name"] = employee_list_df[["First Name"]].copy()
    result["Date Of Birth"] = employee_list_df[["Date Of Birth"]].copy()
    result["Service Start Date"] = employee_list_df[["Service Start Date"]].copy()
    result["Date Of Birth"] = pd.to_datetime(result["Date Of Birth"]).dt.strftime("%d-%m-%Y")
    result["Service Start Date"] = pd.to_datetime(result["Service Start Date"]).dt.strftime("%d-%m-%Y")

    lookup = duckdb.query("""
        SELECT
            el."WorkDay ID",
            ht."Headcount" AS "New Hires Reason"
        FROM employee_list_df el
        LEFT JOIN hire_team_df ht
            ON el."WorkDay ID" = ht."Employee ID"
    """).to_df()
    result = result.merge(lookup, on="WorkDay ID", how="left")

    lookup = duckdb.query("""
        SELECT
            el."WorkDay ID",
            m."[02-BU/CF]" AS "[02-BU/CF]",
            m."[03-SBU/SCF]" AS "[03-SBU/SCF]",
            m."[04-Section]" AS "[04-Section]",
            m."[05-Team]" AS "[05-Team]",
            m."[06-Sub-Team]" AS "[06-Sub-Team]",
            m."[07-Unit]" AS "[07-Unit]",
            m."[08-Sub-Unit]" AS "[08-Sub-Unit]",
            m."[09-Sub - sub unit]" AS "[09-Sub - sub unit]"
        FROM employee_list_df el
        LEFT JOIN new_employee_df ne
            ON TRIM(CAST(el."WorkDay ID" AS VARCHAR)) = TRIM(CAST(ne."WD ID" AS VARCHAR))
        LEFT JOIN master_df m
            ON TRIM(CAST(ne."ID same team" AS VARCHAR)) = TRIM(CAST(m."Employee Local ID" AS VARCHAR))
    """).to_df()
    result = result.merge(lookup, on="WorkDay ID", how="left")

    lookup = duckdb.query("""
        SELECT
            el."WorkDay ID",
            ne."Tỉnh/Thành phố" AS "Tỉnh/Thành phố",
            ne."Xã/Phường" AS "Xã/Phường",
            ne."Địa chỉ đường" AS "Địa chỉ đường",
            ne."Số tài khoản" AS "Số tài khoản",
            ne."Tên ngân hàng" AS "Tên ngân hàng",
            ne."Chi nhánh ngân hàng" AS "Chi nhánh ngân hàng"
        FROM employee_list_df el
        LEFT JOIN new_employee_df ne
            ON TRIM(CAST(el."WorkDay ID" AS VARCHAR)) = TRIM(CAST(ne."WD ID" AS VARCHAR))
    """).to_df()
    result = result.merge(lookup, on="WorkDay ID", how="left")

    result["Tên chủ tài khoản"] = result[["First Name"]].copy()
    result["Tên chủ tài khoản"] = result["Tên chủ tài khoản"].apply(
        lambda x: remove_diacritics(x).upper() if pd.notna(x) else x
    )

    for column in ["Tên ngân hàng", "Tỉnh/Thành phố", "Xã/Phường", "Chi nhánh ngân hàng"]:
        result[column] = result[column].apply(lambda x: remove_diacritics(x) if pd.notna(x) else x)

    llm_model = build_llm_model()
    configs = [
        ColumnConfig(
            name="Tên ngân hàng",
            canonical_list=banks,
            alias_dict=bank_aliases,
            threshold_high=85,
            threshold_fallback=65,
        ),
        ColumnConfig(
            name="Tỉnh/Thành phố",
            canonical_list=provinces,
            alias_dict=province_old_to_new_alias,
            threshold_high=85,
            threshold_fallback=65,
        ),
        ColumnConfig(
            name="Xã/Phường",
            canonical_list=wards,
            alias_dict=ward_old_to_new_alias,
            threshold_high=88,
            threshold_fallback=70,
        ),
    ]

    for config in configs:
        col_values = result[config.name].fillna("").astype(str).tolist()
        matcher = ClosedSetMatcher(config, llm_model=llm_model)
        df_match = matcher.run(col_values)
        result[f"{config.name} (chuẩn hóa)"] = df_match["matched"].values
        result[f"{config.name} (tier)"] = df_match["tier"].values

    new_order = [
        "Employee Local ID",
        "WorkDay ID",
        "First Name",
        "Date Of Birth",
        "Service Start Date",
        "New Hires Reason",
        "[02-BU/CF]",
        "[03-SBU/SCF]",
        "[04-Section]",
        "[05-Team]",
        "[06-Sub-Team]",
        "[07-Unit]",
        "[08-Sub-Unit]",
        "[09-Sub - sub unit]",
        "Tỉnh/Thành phố",
        "Tỉnh/Thành phố (chuẩn hóa)",
        "Tỉnh/Thành phố (tier)",
        "Xã/Phường",
        "Xã/Phường (chuẩn hóa)",
        "Xã/Phường (tier)",
        "Địa chỉ đường",
        "Tên chủ tài khoản",
        "Tên ngân hàng",
        "Tên ngân hàng (chuẩn hóa)",
        "Tên ngân hàng (tier)",
        "Chi nhánh ngân hàng",
        "Số tài khoản",
    ]

    output_path.unlink(missing_ok=True)
    result_with_tier = result[new_order].copy()
    output_cols = [col for col in new_order if not col.endswith(" (tier)")]
    output_df = result_with_tier[output_cols].copy()
    output_df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.worksheets[0]

    red_fill = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid",
    )
    header_to_col = {cell.value: cell.column for cell in ws[1]}

    for config in configs:
        raw_col = config.name
        normalized_col = f"{config.name} (chuẩn hóa)"
        tier_col = f"{config.name} (tier)"

        if raw_col not in header_to_col or normalized_col not in header_to_col:
            continue

        raw_excel_col = header_to_col[raw_col]
        normalized_excel_col = header_to_col[normalized_col]

        for row_idx, tier_value in enumerate(result_with_tier[tier_col], start=2):
            if is_tier2_or_below(tier_value):
                ws.cell(row=row_idx, column=raw_excel_col).fill = red_fill # type: ignore
                ws.cell(row=row_idx, column=normalized_excel_col).fill = red_fill # type: ignore

    for col_idx, col_name in enumerate(output_df.columns, start=1):
        value_lengths = output_df[col_name].map(
            lambda value: len("" if pd.isna(value) else str(value))
        )
        max_len = max(
            value_lengths.max(),
            len(str(col_name)),
        )
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(output_path)
